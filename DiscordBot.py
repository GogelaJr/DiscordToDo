import discord
from discord import File
from io import BytesIO
from discord.ext import commands
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle
import pandas as pd

TOKEN = "Redacted for Security Purposes"

client = commands.Bot(command_prefix='!')

thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def create_worksheet():
    if os.path.exists('todo_list.xlsx'):
        workbook = openpyxl.load_workbook('todo_list.xlsx')
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "To-Do List"
        header = ["ID", "TASK", "PRIORITY", "CREATED", "COMPLETED", "TIME ELAPSED", "STATUS", "CREATED BY"]

        sheet.append(header)
        

        for cell in header:
            cell_obj = sheet.cell(1, header.index(cell) + 1)
            cell_obj.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_obj.border = thick_border
            cell_obj.font = Font(name="Sylfaen", size=14, bold=True, color="FFFFFF")
            cell_obj.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 10)
            sheet.column_dimensions[column].width = adjusted_width
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20

    return workbook, sheet

def create_messagesheet():
    if os.path.exists('message.xlsx'):
        workbook = openpyxl.load_workbook('message.xlsx')
        sheet = workbook.active
    else:
        messagebook = openpyxl.Workbook()
        messagesheet = messagebook.active
        messagesheet.title = "Message List"
        header = ["ID", "AUTHOR", "MESSAGE", "SENT ON"]

        messagesheet.append(header)
        

        for cell in header:
            cell_obj = messagesheet.cell(1, header.index(cell) + 1)
            cell_obj.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_obj.border = thick_border
            
            cell_obj.font = Font(name="Sylfaen", size=14, bold=True, color="FFFFFF")
            cell_obj.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        for col in messagesheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 10)
            messagesheet.column_dimensions[column].width = adjusted_width
        messagesheet.column_dimensions['B'].width = 20
        messagesheet.column_dimensions['C'].width = 20
        messagesheet.column_dimensions['D'].width = 20

    return messagebook, messagesheet


# save the workbook
workbook, sheet = create_worksheet()
workbook.save("todo_list.xlsx")

messagebook, messagesheet = create_messagesheet()
messagebook.save("messages.xlsx")





def current_time():

    current_date = datetime.today()
    date_string = current_date.strftime('%d/%b/%Y %H:%M:%S')
    return date_string

from datetime import datetime, timedelta

def time_elapsed(created_time):
    current_time = datetime.now()
    date_format = '%d/%b/%Y %H:%M:%S'
    date_time_obj = datetime.strptime(created_time, date_format)
    elapsed_time = current_time - date_time_obj
    days = elapsed_time.days
    hours = elapsed_time.seconds // 3600
    minutes = (elapsed_time.seconds % 3600) // 60
    seconds = elapsed_time.seconds % 60

    elapsed_days = 0
    elapsed_months = 0
    elapsed_years = 0

    if days >= 365:
        elapsed_years = days // 365
        days = days % 365
    if days >= 30:
        elapsed_months = days // 30
        days = days % 30
    if days > 0:
        elapsed_days = days
        return

    return f"{hours}:" \
           f"{minutes}:" \
           f"{seconds}" \


def get_row_by_id(tid):
    row = df.loc[df['ID'] == tid]
    if len(row) == 0:
        return None
    else:
        return row

def read_tasks_from_excel():
    df = pd.read_excel('todo_list.xlsx', engine="openpyxl")
    return df


def reapplyStyles(filepath):
    workbook = openpyxl.load_workbook("todo_list.xlsx")
    sheet = workbook.active
    font = Font(name="Sylfaen", size=14, bold=True, color="FFFFFF")
    fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    border = thick_border
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in sheet[1]:
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = align
        sheet.column_dimensions[cell.column_letter].width = 20
    workbook.save("todo_list.xlsx")


#Create new Task
@client.command()
async def addtask(ctx, tname, priority):
    last_row = sheet.max_row
    print(f'{ctx.author} added new TID: {last_row} Task: {tname}, Priority: {priority} on {current_time()}')
    sheet.append([last_row, tname, priority, current_time(), " ", " ", "Incomplete", str(ctx.author)])
    workbook.save("todo_list.xlsx")
    await ctx.send(f'{ctx.author.mention}\nTID: {last_row}\nTask: {tname}\nPriority: {priority}\nhas been added to the Tasks')


#Task Details
@client.command()
async def details(ctx, tid: int):
    df = read_tasks_from_excel()
    # Find the task with the specified ID
    task = df.loc[df['ID'] == tid]
    # Check if the task was found
    if len(task) == 0:
        await ctx.send(f"No task found with ID {tid}")
    else:
        task_name = task.iloc[0]['TASK']
        priority = task.iloc[0]['PRIORITY']
        date = task.iloc[0]['CREATED']  
        createdby = task.iloc[0]["CREATED BY"]
        completed = task.iloc[0]["COMPLETED"]
        time_elapsed = task.iloc[0]["TIME ELAPSED"]
        status = task.iloc[0]["STATUS"]
        await ctx.send(f"TID: {tid}\nTask Name: {task_name}\nPriority: {priority}\nCreated on: {date}\nCreated By: {createdby}\nStatus: {status}\nTime Elapsed: {time_elapsed}\nCompleted on: {completed}")


# Status for Tasks
@client.command()
async def finish(ctx, tid:int):
    df = read_tasks_from_excel()
    task = df.loc[df['ID'] == tid]
    if not task.empty:
        if task.iloc[0]["STATUS"] == "Finished":
            await ctx.send(f'{ctx.author.mention} Task #{tid} has been already completed on {task.iloc[0]["COMPLETED"]}')
        else:
            df.at[task.index[0], 'STATUS'] = "Finished"
            df.at[task.index[0], 'COMPLETED'] = current_time()
            df.at[task.index[0], 'TIME ELAPSED'] = time_elapsed(df.at[task.index[0], 'CREATED'])
            df.to_excel("todo_list.xlsx", index=False)
            reapplyStyles("todo_list.xlsx")
            print(f'{ctx.author} completed Task #{tid}')
            await ctx.send(f'{ctx.author.mention} Task #{tid} has been completed on {current_time()}')
    else:
        await ctx.send(f'{ctx.author.mention} Task #{tid} not found')


@client.command()
async def inprogress(ctx, tid:int):
    
    df = read_tasks_from_excel()
    task = df.loc[df['ID'] == tid]
    if not task.empty:
        df.at[task.index[0], 'STATUS'] = "In Progress"
        df.to_excel("todo_list.xlsx", index=False)
        print(f'{ctx.author} completed Task #{tid}')
        reapplyStyles("todo_list.xlsx")
        await ctx.send(f'{ctx.author.mention} Task #{tid} is in Progress')
    else:
        await ctx.send(f'{ctx.author.mention} Task #{tid} not found')

@client.command()
async def commands(ctx):
    await ctx.send("""To use the commands make sure to use the prefix !
1)  addtask Task Priority - to create a new Task.
2)  details TID - to see the details of the Task.
3)  finish TID - to change the status of the Task to Finished.
4)  inprogress TID - to change the status of the Task to In Progress.
5)  exportlist - to export the current Task List""")


@client.command()
async def exportlist(ctx):
    wb = openpyxl.load_workbook("todo_list.xlsx")
    ws = wb.active

    # Save the Excel file to a byte stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Send the Excel file as an attachment
    file = discord.File(stream, filename="todo_list.xlsx")
    await ctx.send(file=file)


@client.command()
async def totaltasks(ctx):
    df = read_tasks_from_excel()
    incompleted = len(df[df['STATUS'] == "Incomplete"])
    finished = len(df[df['STATUS'] == "Finished"])
    inprogress = len(df[df['STATUS'] == "In Progress"])
    await ctx.send(f'{ctx.author.mention}\nRegistered: {incompleted}\nIn Progress Tasks: {inprogress}\nFinished: {finished}')
## Admin Message in case Urgency or Bug Reporting
@client.command()
async def messageme(ctx, message):
    last_row = messagesheet.max_row
    print(f'{current_time()} {ctx.author}: {message}')
    messagesheet.append([last_row, str(ctx.author), message, current_time()])
    messagebook.save("messages.xlsx")
    await ctx.send(f'{ctx.author.mention} message delivered!')

client.run(TOKEN)
